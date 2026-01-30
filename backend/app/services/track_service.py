"""
轨迹服务层
"""
import gpxpy
import asyncio
from datetime import datetime, timezone
from typing import Optional, List
from sqlalchemy import select, func, delete, insert, and_, update
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.track import Track, TrackPoint
from app.models.user import User
from app.gpxutil_wrapper.coord_transform import convert_point_to_all, CoordinateType
from app.gpxutil_wrapper.geocoding import create_geocoding_service, GeocodingProvider
from app.core.query_helper import QueryHelper, SoftDeleteMixin, AuditMixin as QueryAuditMixin
from loguru import logger


class TrackService:
    """轨迹服务类"""

    def __init__(self):
        self._geocoding_service = None
        self._geocoding_provider = None
        self._geocoding_config = None

        # 存储当前正在填充的轨迹进度
        self._filling_progress = {}  # {track_id: {"current": 10, "total": 100, "status": "filling"}}

        # 追踪后台任务
        self._background_tasks = set()  # type: set[asyncio.Task]

    async def cancel_all_tasks(self):
        """取消所有后台任务"""
        if not self._background_tasks:
            return

        # 复制任务列表，避免迭代过程中被修改
        tasks_to_cancel = list(self._background_tasks)

        # 取消所有未完成的任务
        for task in tasks_to_cancel:
            if not task.done():
                task.cancel()

        # 逐个等待任务完成（避免 gather 的问题）
        import logging
        logger = logging.getLogger(__name__)

        for task in tasks_to_cancel:
            try:
                # 使用 wait_for 避免无限等待
                await asyncio.wait_for(task, timeout=0.5)
            except asyncio.CancelledError:
                # 任务被取消，这是预期的
                pass
            except asyncio.TimeoutError:
                # 超时，继续下一个任务
                logger.warning(f"Task {task.get_name()} did not cancel in time")
            except Exception as e:
                # 其他异常，记录但继续
                logger.warning(f"Error while cancelling task {task.get_name()}: {e}")

        # 清空任务列表
        self._background_tasks.clear()

    async def _get_geocoding_service(self, db: AsyncSession):
        """获取地理编码服务实例"""
        from app.services.config_service import config_service

        provider = await config_service.get(db, "geocoding_provider")
        if not provider:
            return None

        if self._geocoding_service is None or self._geocoding_provider != provider:
            self._geocoding_provider = provider
            config = await config_service.get_json(db, "geocoding_config", {})
            provider_config = config.get(provider, {})
            self._geocoding_service = create_geocoding_service(provider, provider_config)

        return self._geocoding_service

    async def get_by_id(self, db: AsyncSession, track_id: int, user_id: int, load_recording: bool = False) -> Optional[Track]:
        """
        获取轨迹（带用户权限检查）

        Args:
            db: 数据库会话
            track_id: 轨迹ID
            user_id: 用户ID
            load_recording: 是否预加载关联的实时记录

        Returns:
            轨迹对象，如果不存在则返回 None
        """
        query = select(Track).where(
            and_(
                Track.id == track_id,
                Track.user_id == user_id,
                Track.is_valid == True
            )
        )

        # 如果需要预加载关联的实时记录
        if load_recording:
            query = query.options(selectinload(Track.live_recordings))

        result = await db.execute(query)
        return result.scalar_one_or_none()

    async def get_by_id_no_check(self, db: AsyncSession, track_id: int) -> Optional[Track]:
        """获取轨迹（不带权限检查）"""
        result = await db.execute(
            select(Track).where(and_(Track.id == track_id, Track.is_valid == True))
        )
        return result.scalar_one_or_none()

    async def get_list(
        self,
        db: AsyncSession,
        user_id: int,
        skip: int = 0,
        limit: int = 20,
        search: Optional[str] = None,
        sort_by: Optional[str] = None,
        sort_order: Optional[str] = "desc",
    ) -> tuple[List[Track], int]:
        """
        获取用户的轨迹列表

        Args:
            db: 数据库会话
            user_id: 用户ID
            skip: 跳过记录数
            limit: 返回记录数
            search: 搜索轨迹名称（模糊匹配）
            sort_by: 排序字段 (start_time, distance, duration)
            sort_order: 排序方向 (asc=正序, desc=倒序)

        Returns:
            (轨迹列表, 总数)
        """
        # 构建基础查询条件
        conditions = [Track.user_id == user_id, Track.is_valid == True]

        # 添加搜索条件
        if search:
            conditions.append(Track.name.ilike(f"%{search}%"))

        # 获取总数
        count_result = await db.execute(
            select(func.count(Track.id)).where(and_(*conditions))
        )
        total = count_result.scalar() or 0

        # 确定排序字段和方向
        sort_field = Track.start_time  # 默认按轨迹开始时间排序
        if sort_by == "distance":
            sort_field = Track.distance
        elif sort_by == "duration":
            sort_field = Track.duration

        # 排序方向
        if sort_order == "asc":
            order_by_clause = sort_field.asc()
        else:
            order_by_clause = sort_field.desc()

        # 获取列表
        result = await db.execute(
            select(Track)
            .where(and_(*conditions))
            .order_by(order_by_clause)
            .offset(skip)
            .limit(limit)
        )
        tracks = list(result.scalars().all())

        return tracks, total

    def _calculate_speed(self, current_time, prev_time, distance_m: float) -> Optional[float]:
        """
        计算速度 (m/s)

        Args:
            current_time: 当前点的时间
            prev_time: 前一个点的时间
            distance_m: 两点之间的距离（米）

        Returns:
            速度（米/秒），如果无法计算则返回 None
        """
        if not current_time or not prev_time:
            return None

        time_diff = (current_time - prev_time).total_seconds()
        if time_diff <= 0:
            return None

        speed = distance_m / time_diff
        # 如果速度异常（超过 200 km/h），返回 None
        if speed > 55.56:  # 200 km/h = 55.56 m/s
            return None

        return round(speed, 2)

    def _calculate_bearing(self, lat1: float, lon1: float, lat2: float, lon2: float) -> Optional[float]:
        """
        计算从点1到点2的方位角（度）

        Args:
            lat1, lon1: 第一个点的纬度和经度（度）
            lat2, lon2: 第二个点的纬度和经度（度）

        Returns:
            方位角（度），范围 [0, 360)，如果两点重合则返回 None
        """
        from math import radians, degrees, atan2, sin, cos, pi

        # 转换为弧度
        lat1_rad = radians(lat1)
        lon1_rad = radians(lon1)
        lat2_rad = radians(lat2)
        lon2_rad = radians(lon2)

        dlon = lon2_rad - lon1_rad

        # 计算方位角
        x = sin(dlon) * cos(lat2_rad)
        y = cos(lat1_rad) * sin(lat2_rad) - sin(lat1_rad) * cos(lat2_rad) * cos(dlon)

        bearing = atan2(x, y)
        bearing_deg = degrees(bearing)

        # 转换到 [0, 360) 范围
        if bearing_deg < 0:
            bearing_deg += 360

        # 检查两点是否重合（方位角无意义）
        if abs(lat1 - lat2) < 1e-9 and abs(lon1 - lon2) < 1e-9:
            return None

        return round(bearing_deg, 2)

    async def create_from_gpx(
        self,
        db: AsyncSession,
        user: User,
        filename: str,
        gpx_content: str,
        name: str,
        description: Optional[str] = None,
        original_crs: CoordinateType = 'wgs84',
        convert_to: Optional[CoordinateType] = None,
        fill_geocoding: bool = False,
    ) -> Track:
        """
        从 GPX 内容创建轨迹

        Args:
            db: 数据库会话
            user: 用户对象
            filename: 原始文件名
            gpx_content: GPX 文件内容
            name: 轨迹名称
            description: 轨迹描述
            original_crs: 原始坐标系
            convert_to: 转换到目标坐标系
            fill_geocoding: 是否填充行政区划和道路信息

        Returns:
            创建的轨迹对象
        """
        # 解析 GPX
        gpx = gpxpy.parse(gpx_content)

        if not gpx.tracks:
            raise ValueError("GPX 文件中没有轨迹数据")

        # 获取第一个轨迹的第一段
        track = gpx.tracks[0]
        segment = track.segments[0]

        if not segment.points:
            raise ValueError("GPX 文件中没有轨迹点")

        # 计算统计信息
        total_distance = 0
        elevation_gain = 0
        elevation_loss = 0
        prev_elevation = None
        start_time = None
        end_time = None

        points_data = []
        prev_point_data = None

        for idx, point in enumerate(segment.points):
            # 时间
            if point.time:
                if start_time is None:
                    start_time = point.time
                end_time = point.time

            # 坐标转换
            all_coords = convert_point_to_all(
                point.longitude,
                point.latitude,
                original_crs
            )

            # 海拔变化
            if point.elevation is not None:
                if prev_elevation is not None:
                    diff = point.elevation - prev_elevation
                    if diff > 0:
                        elevation_gain += diff
                    else:
                        elevation_loss += abs(diff)
                prev_elevation = point.elevation

            # 计算距离和速度
            distance_from_prev = 0
            speed = None
            if idx > 0 and prev_point_data:
                # 使用 WGS84 坐标计算 3D 距离
                from math import sqrt, radians, sin, cos, atan2

                lat1 = radians(prev_point_data['latitude_wgs84'])
                lon1 = radians(prev_point_data['longitude_wgs84'])
                lat2 = radians(all_coords['wgs84'][1])
                lon2 = radians(all_coords['wgs84'][0])

                # Haversine 公式计算水平距离
                dlat = lat2 - lat1
                dlon = lon2 - lon1
                a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
                c = 2 * atan2(sqrt(a), sqrt(1 - a))
                horizontal_distance = 6371000 * c  # 地球半径 6371km

                # 计算垂直距离
                elev1 = prev_point_data.get('elevation') or 0
                elev2 = point.elevation or 0
                vertical_distance = elev2 - elev1

                # 3D 距离
                distance_from_prev = sqrt(horizontal_distance ** 2 + vertical_distance ** 2)
                total_distance += distance_from_prev

                # 计算速度
                speed = self._calculate_speed(point.time, prev_point_data['time'], distance_from_prev)

                # 计算方位角
                bearing = self._calculate_bearing(
                    prev_point_data['latitude_wgs84'],
                    prev_point_data['longitude_wgs84'],
                    all_coords['wgs84'][1],
                    all_coords['wgs84'][0]
                )

            point_data = {
                'point_index': idx,
                'time': point.time,
                'latitude_wgs84': all_coords['wgs84'][1],
                'longitude_wgs84': all_coords['wgs84'][0],
                'latitude_gcj02': all_coords['gcj02'][1],
                'longitude_gcj02': all_coords['gcj02'][0],
                'latitude_bd09': all_coords['bd09'][1],
                'longitude_bd09': all_coords['bd09'][0],
                'elevation': point.elevation,
                'speed': speed,
                'bearing': bearing if idx > 0 else None,  # 第一个点没有方位角
            }
            points_data.append(point_data)
            prev_point_data = point_data

        # 计算时长（秒）
        duration = 0
        if start_time and end_time:
            duration = int((end_time - start_time).total_seconds())

        # 创建轨迹记录，带审计字段
        track_obj = Track(
            user_id=user.id,
            name=name,
            description=description,
            original_filename=filename,
            original_crs=original_crs,
            distance=round(total_distance, 2),
            duration=duration,
            elevation_gain=round(elevation_gain, 2),
            elevation_loss=round(elevation_loss, 2),
            start_time=start_time,
            end_time=end_time,
            has_area_info=False,
            has_road_info=False,
            created_by=user.id,
            updated_by=user.id,
            is_valid=True,
        )
        db.add(track_obj)
        await db.flush()  # 获取 track_id

        # 批量创建轨迹点
        insert_values = []
        for point_data in points_data:
            insert_values.append({
                "track_id": track_obj.id,
                "point_index": point_data["point_index"],
                "time": point_data["time"],
                "latitude_wgs84": point_data["latitude_wgs84"],
                "longitude_wgs84": point_data["longitude_wgs84"],
                "latitude_gcj02": point_data["latitude_gcj02"],
                "longitude_gcj02": point_data["longitude_gcj02"],
                "latitude_bd09": point_data["latitude_bd09"],
                "longitude_bd09": point_data["longitude_bd09"],
                "elevation": point_data["elevation"],
                "speed": point_data["speed"],
                "bearing": point_data["bearing"],
                "province": None,
                "city": None,
                "district": None,
                "province_en": None,
                "city_en": None,
                "district_en": None,
                "road_name": None,
                "road_number": None,
                "road_name_en": None,
                "created_by": user.id,
                "updated_by": user.id,
                "is_valid": True,
            })

        # 分批插入，每批 500 条
        batch_size = 500
        for i in range(0, len(insert_values), batch_size):
            batch = insert_values[i:i + batch_size]
            try:
                await db.execute(
                    TrackPoint.__table__.insert().values(batch)
                )
            except Exception as e:
                logger.error(f"Error inserting batch {i//batch_size}: {e}")
                raise

        await db.commit()
        await db.refresh(track_obj)

        # 异步填充地理信息（如果启用）
        if fill_geocoding:
            task = asyncio.create_task(
                self.fill_geocoding_info(db, track_obj.id, user.id)
            )
            task.add_done_callback(self._background_tasks.discard)
            self._background_tasks.add(task)

        return track_obj

    def get_fill_progress(self, track_id: int) -> dict:
        """获取地理信息填充进度"""
        return self._filling_progress.get(track_id, {"current": 0, "total": 0, "status": "idle"})

    def stop_fill_geocoding(self, track_id: int) -> bool:
        """停止地理信息填充"""
        if track_id in self._filling_progress:
            progress = self._filling_progress[track_id]
            if progress.get("status") == "filling":
                progress["status"] = "stopped"
                return True
        return False

    async def fill_geocoding_info(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
    ):
        """
        填充行政区划和道路信息（合并功能）

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID
        """
        # 使用新的会话，因为原会话可能已关闭
        from app.core.database import async_session_maker

        logger.info(f"Starting geocoding fill for track {track_id}")

        async with async_session_maker() as db:
            try:
                # 获取轨迹点（按时间排序，实时记录场景下 point_index 可能乱序）
                result = await db.execute(
                    select(TrackPoint)
                    .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
                    .order_by(TrackPoint.time, TrackPoint.created_at)
                )
                points = result.scalars().all()

                if not points:
                    logger.warning(f"No points found for track {track_id}")
                    return

                total_points = len(points)
                logger.info(f"Found {total_points} points for track {track_id}")

                # 初始化进度
                self._filling_progress[track_id] = {
                    "current": 0,
                    "total": total_points,
                    "status": "filling"
                }

                # 获取地理编码服务
                geocoding_service = await self._get_geocoding_service(db)
                if not geocoding_service:
                    logger.warning(f"Geocoding service not configured for track {track_id}")
                    self._filling_progress[track_id]["status"] = "failed"
                    self._filling_progress[track_id]["error"] = "Geocoding service not configured"
                    return

                logger.info(f"Geocoding service acquired: {type(geocoding_service).__name__}")

                updated_count = 0
                for idx, point in enumerate(points):
                    # 检查是否已停止
                    if self._filling_progress[track_id].get("status") == "stopped":
                        logger.info(f"Geocoding fill stopped for track {track_id}")
                        return

                    # 检查任务是否被取消
                    try:
                        # 使用 asyncio.sleep 的同时检查取消
                        await asyncio.sleep(0)
                    except asyncio.CancelledError:
                        logger.info(f"Geocoding fill cancelled for track {track_id}")
                        self._filling_progress[track_id]["status"] = "stopped"
                        return

                    try:
                        lat = point.latitude_wgs84
                        lon = point.longitude_wgs84
                        info = await geocoding_service.get_point_info(lat, lon)

                        # 同时填充行政区划和道路信息
                        point.province = info.get('province', '')
                        point.city = info.get('city', '')
                        point.district = info.get('area', '')
                        point.road_name = info.get('road_name', '')
                        point.road_number = info.get('road_num', '')
                        # 英文字段
                        point.province_en = info.get('province_en', '')
                        point.city_en = info.get('city_en', '')
                        point.district_en = info.get('area_en', '')
                        point.road_name_en = info.get('road_name_en', '')

                        # 更新审计字段
                        point.updated_by = user_id

                        updated_count += 1

                        # 更新进度
                        self._filling_progress[track_id]["current"] = idx + 1

                        # 每次请求后短暂延迟，避免过载
                        try:
                            await asyncio.sleep(0.2)
                        except asyncio.CancelledError:
                            logger.info(f"Geocoding fill cancelled for track {track_id} (during sleep)")
                            self._filling_progress[track_id]["status"] = "stopped"
                            raise  # 重新抛出，让上层处理

                    except asyncio.CancelledError:
                        # 任务被取消，标记为已停止并重新抛出
                        logger.info(f"Geocoding fill cancelled for track {track_id}")
                        self._filling_progress[track_id]["status"] = "stopped"
                        raise
                    except Exception as e:
                        logger.error(f"Error getting geocoding for point {idx}: {e}")

                # 更新轨迹标记
                track_result = await db.execute(
                    select(Track).where(Track.id == track_id)
                )
                track = track_result.scalar_one_or_none()
                if track:
                    track.has_area_info = True
                    track.has_road_info = True
                    track.updated_by = user_id

                await db.commit()

                # 标记完成
                self._filling_progress[track_id]["status"] = "completed"
                logger.info(f"Geocoding info filled for track {track_id}, updated {updated_count} points")

            except Exception as e:
                logger.error(f"Error filling geocoding info for track {track_id}: {e}")
                self._filling_progress[track_id]["status"] = "failed"
                self._filling_progress[track_id]["error"] = str(e)
                await db.rollback()

    async def update(self, db: AsyncSession, track: Track, update_data, user_id: int) -> Track:
        """更新轨迹信息（名称和描述）"""
        update_dict = update_data.model_dump(exclude_unset=True)

        if update_dict:
            for field, value in update_dict.items():
                setattr(track, field, value)

            track.updated_by = user_id
            track.updated_at = datetime.now(timezone.utc)

            await db.commit()
            await db.refresh(track)

        return track

    async def delete(self, db: AsyncSession, track: Track, user_id: int) -> None:
        """软删除轨迹"""
        track.is_valid = False
        track.updated_by = user_id
        track.updated_at = datetime.now(timezone.utc)

        # 同时软删除所有轨迹点
        await db.execute(
            update(TrackPoint)
            .where(TrackPoint.track_id == track.id)
            .values(
                is_valid=False,
                updated_by=user_id,
                updated_at=datetime.now(timezone.utc)
            )
        )

        await db.commit()

    async def get_stats(self, db: AsyncSession, user_id: int) -> dict:
        """获取用户的轨迹统计"""
        result = await db.execute(
            select(
                func.count(Track.id).label('total_tracks'),
                func.sum(Track.distance).label('total_distance'),
                func.sum(Track.duration).label('total_duration'),
                func.sum(Track.elevation_gain).label('total_elevation_gain'),
            ).where(and_(Track.user_id == user_id, Track.is_valid == True))
        )
        row = result.one()

        return {
            'total_tracks': row.total_tracks or 0,
            'total_distance': row.total_distance or 0,
            'total_duration': row.total_duration or 0,
            'total_elevation_gain': row.total_elevation_gain or 0,
        }

    async def get_live_recording_stats(self, db: AsyncSession, user_id: int) -> dict:
        """获取用户实时记录的统计（从关联的当前轨迹获取）"""
        from app.models.live_recording import LiveRecording

        # 获取所有实时记录
        result = await db.execute(
            select(LiveRecording).where(
                and_(
                    LiveRecording.user_id == user_id,
                    LiveRecording.is_valid == True
                )
            )
        )
        recordings = list(result.scalars().all())

        if not recordings:
            return {
                'total_tracks': 0,
                'total_distance': 0,
                'total_duration': 0,
                'total_elevation_gain': 0,
            }

        # 获取所有关联的当前轨迹 ID（只统计已有关联轨迹的实时记录）
        track_ids = [r.current_track_id for r in recordings if r.current_track_id]

        if not track_ids:
            return {
                'total_tracks': 0,  # 没有关联轨迹的实时记录不计入统计
                'total_distance': 0,
                'total_duration': 0,
                'total_elevation_gain': 0,
            }

        # 获取这些轨迹的统计
        result = await db.execute(
            select(
                func.count(Track.id).label('total_tracks'),
                func.sum(Track.distance).label('total_distance'),
                func.sum(Track.duration).label('total_duration'),
                func.sum(Track.elevation_gain).label('total_elevation_gain'),
            ).where(
                and_(
                    Track.id.in_(track_ids),
                    Track.is_valid == True
                )
            )
        )
        row = result.one()

        return {
            'total_tracks': row.total_tracks or 0,  # 使用实际轨迹数量
            'total_distance': row.total_distance or 0,
            'total_duration': row.total_duration or 0,
            'total_elevation_gain': row.total_elevation_gain or 0,
        }

    async def get_unified_list(
        self,
        db: AsyncSession,
        user_id: int,
        skip: int = 0,
        limit: int = 20,
        search: Optional[str] = None,
        sort_by: Optional[str] = None,
        sort_order: Optional[str] = "desc",
    ) -> tuple[List[dict], int]:
        """
        获取用户的轨迹列表（包含实时记录状态）

        返回普通轨迹列表，如果轨迹有正在进行的实时记录，会添加 recording_status 字段
        正在记录的轨迹会优先显示在列表前面
        没有关联轨迹的实时记录也会显示（显示记录名称，数据为0）

        Returns:
            (轨迹列表, 总数)
        """
        from app.models.live_recording import LiveRecording

        # 构建查询条件
        track_conditions = [Track.user_id == user_id, Track.is_valid == True]
        if search:
            track_conditions.append(Track.name.ilike(f"%{search}%"))

        # 获取轨迹总数
        track_count_result = await db.execute(
            select(func.count(Track.id)).where(and_(*track_conditions))
        )
        track_total = track_count_result.scalar() or 0

        # 获取所有正在进行的实时记录
        recording_conditions = [
            LiveRecording.user_id == user_id,
            LiveRecording.is_valid == True,
            LiveRecording.status == 'active',
        ]
        if search:
            recording_conditions.append(LiveRecording.name.ilike(f"%{search}%"))

        recording_count_result = await db.execute(
            select(func.count(LiveRecording.id)).where(and_(*recording_conditions))
        )
        recording_total = recording_count_result.scalar() or 0

        # 总数 = 普通轨迹 + 正在记录的实时记录（去重）
        # 但需要排除已关联轨迹的实时记录（避免重复计数）
        # 先获取已关联轨迹的实时记录数量
        linked_recording_count_result = await db.execute(
            select(func.count(LiveRecording.id)).where(
                and_(
                    LiveRecording.user_id == user_id,
                    LiveRecording.is_valid == True,
                    LiveRecording.status == 'active',
                    LiveRecording.current_track_id.isnot(None)
                )
            )
        )
        linked_recording_count = linked_recording_count_result.scalar() or 0

        # 计算去重后的总数：普通轨迹 + 未关联的实时记录
        total = track_total + (recording_total - linked_recording_count)

        # 获取所有轨迹
        sort_field = Track.created_at
        if sort_by == "start_time":
            sort_field = Track.start_time
        elif sort_by == "distance":
            sort_field = Track.distance
        elif sort_by == "duration":
            sort_field = Track.duration

        if sort_order == "asc":
            order_by_clause = sort_field.asc()
        else:
            order_by_clause = sort_field.desc()

        track_result = await db.execute(
            select(Track)
            .where(and_(*track_conditions))
            .order_by(order_by_clause)
        )
        all_tracks = list(track_result.scalars().all())

        # 获取所有正在进行的实时记录
        recording_result = await db.execute(
            select(LiveRecording)
            .where(and_(*recording_conditions))
            .order_by(LiveRecording.created_at.desc())
        )
        all_recordings = list(recording_result.scalars().all())

        # 构建 track_id -> recording 的映射
        recording_map = {r.current_track_id: r for r in all_recordings if r.current_track_id}

        # 转换为响应格式
        unified_items = []
        recording_tracks = []
        normal_tracks = []

        # 处理普通轨迹
        for track in all_tracks:
            recording = recording_map.get(track.id)

            # 获取 last_point_time（如果有实时记录）
            last_point_time = None
            if recording:
                from app.services.live_recording_service import live_recording_service
                last_point_time = await live_recording_service.get_last_point_time(db, recording)

            item = {
                'id': track.id,
                'user_id': track.user_id,
                'name': track.name,
                'description': track.description,
                'original_filename': track.original_filename,
                'original_crs': track.original_crs,
                'distance': track.distance,
                'duration': track.duration,
                'elevation_gain': track.elevation_gain,
                'elevation_loss': track.elevation_loss,
                'start_time': track.start_time,
                'end_time': track.end_time,
                'has_area_info': track.has_area_info,
                'has_road_info': track.has_road_info,
                'created_at': track.created_at,
                'updated_at': track.updated_at,
                # 实时记录状态字段
                'is_live_recording': recording is not None,
                'live_recording_id': recording.id if recording else None,
                'live_recording_status': recording.status if recording else None,
                'live_recording_token': recording.token if recording else None,
                'fill_geocoding': recording.fill_geocoding if recording else False,
                'last_upload_at': recording.last_upload_at if recording else None,
                'last_point_time': last_point_time,
            }

            if recording:
                recording_tracks.append(item)
            else:
                normal_tracks.append(item)

        # 处理没有关联轨迹的实时记录（创建虚拟轨迹项）
        linked_track_ids = {r.current_track_id for r in all_recordings if r.current_track_id}
        for recording in all_recordings:
            if recording.current_track_id is None:
                # 没有关联轨迹，创建虚拟项
                recording_tracks.append({
                    'id': -recording.id,  # 使用负数 ID 表示这是虚拟项（前端用于判断）
                    'user_id': recording.user_id,
                    'name': recording.name,
                    'description': recording.description,
                    'original_filename': None,
                    'original_crs': 'wgs84',
                    'distance': 0,
                    'duration': 0,
                    'elevation_gain': 0,
                    'elevation_loss': 0,
                    'start_time': recording.created_at,
                    'end_time': recording.last_upload_at,
                    'has_area_info': False,
                    'has_road_info': False,
                    'created_at': recording.created_at,
                    'updated_at': recording.updated_at,
                    # 实时记录状态字段
                    'is_live_recording': True,
                    'live_recording_id': recording.id,
                    'live_recording_status': recording.status,
                    'live_recording_token': recording.token,
                    'fill_geocoding': recording.fill_geocoding or False,
                    'last_upload_at': recording.last_upload_at,
                    'last_point_time': None,  # 没有关联轨迹，无法获取轨迹点时间
                })

        # 合并：正在记录的轨迹在前，普通轨迹在后
        unified_items = recording_tracks + normal_tracks

        # 应用分页
        items = unified_items[skip:skip + limit]

        return items, total

    async def get_points(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
        crs: CoordinateType = 'wgs84',
    ) -> List[TrackPoint]:
        """
        获取轨迹点数据

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID
            crs: 返回的坐标系

        Returns:
            轨迹点列表
        """
        # 检查权限
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            return []

        # 获取轨迹点（按时间排序）
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.time, TrackPoint.created_at)
        )
        return list(result.scalars().all())

    async def update_bearings_for_track(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
    ) -> dict:
        """
        为指定轨迹的所有点计算并更新方位角

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID

        Returns:
            更新结果 {"updated": 更新的点数, "total": 总点数}
        """
        # 检查权限
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            raise ValueError("轨迹不存在")

        # 获取所有轨迹点（按时间排序）
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.time, TrackPoint.created_at)
        )
        points = list(result.scalars().all())

        if not points:
            return {"updated": 0, "total": 0}

        updated_count = 0
        prev_point = None

        for point in points:
            # 第一个点没有方位角
            if prev_point is None:
                point.bearing = None
            else:
                # 计算从前一个点到当前点的方位角
                bearing = self._calculate_bearing(
                    prev_point.latitude_wgs84,
                    prev_point.longitude_wgs84,
                    point.latitude_wgs84,
                    point.longitude_wgs84
                )
                point.bearing = bearing
                updated_count += 1

            point.updated_by = user_id
            prev_point = point

        await db.commit()

        logger.info(f"Updated bearings for track {track_id}: {updated_count}/{len(points)} points")

        return {"updated": updated_count, "total": len(points)}

    async def update_bearings_for_all_tracks(
        self,
        db: AsyncSession,
    ) -> dict:
        """
        为所有现有轨迹计算并更新方位角

        Args:
            db: 数据库会话

        Returns:
            更新结果 {"updated_tracks": 更新的轨迹数, "updated_points": 更新的点数, "total_tracks": 总轨迹数}
        """
        # 获取所有有效轨迹
        result = await db.execute(
            select(Track).where(Track.is_valid == True)
        )
        tracks = list(result.scalars().all())

        updated_tracks = 0
        updated_points = 0

        for track in tracks:
            try:
                # 获取轨迹的所有点（按时间排序）
                points_result = await db.execute(
                    select(TrackPoint)
                    .where(and_(TrackPoint.track_id == track.id, TrackPoint.is_valid == True))
                    .order_by(TrackPoint.time, TrackPoint.created_at)
                )
                points = list(points_result.scalars().all())

                if not points:
                    continue

                prev_point = None
                for point in points:
                    if prev_point is None:
                        point.bearing = None
                    else:
                        bearing = self._calculate_bearing(
                            prev_point.latitude_wgs84,
                            prev_point.longitude_wgs84,
                            point.latitude_wgs84,
                            point.longitude_wgs84
                        )
                        point.bearing = bearing
                        updated_points += 1

                    point.updated_by = track.user_id
                    prev_point = point

                updated_tracks += 1

                # 每处理一个轨迹提交一次
                await db.commit()

                logger.info(f"Updated bearings for track {track.id} ({track.name})")

            except Exception as e:
                logger.error(f"Error updating bearings for track {track.id}: {e}")
                await db.rollback()

        logger.info(f"Bearing update completed: {updated_tracks} tracks, {updated_points} points")

        return {
            "updated_tracks": updated_tracks,
            "updated_points": updated_points,
            "total_tracks": len(tracks)
        }

    async def get_timeline(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
    ) -> List[dict]:
        """
        获取轨迹的区域时间线

        按时间顺序聚合轨迹点，当区域发生变化时创建新的时间线条目。
        区域包括：省、市、区、道路名称、道路编号

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID

        Returns:
            时间线条目列表，每个条目包含区域信息和时间范围
        """
        # 检查权限
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            return []

        # 获取轨迹点（按时间排序）
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.time, TrackPoint.created_at)
        )
        points = list(result.scalars().all())

        if not points:
            return []

        timeline_entries = []
        current_entry = None
        point_count = 0

        for point in points:
            # 构建区域标识符（用于检测区域变化）
            region_key = (
                point.province or '',
                point.city or '',
                point.district or '',
                point.road_name or '',
                point.road_number or '',
            )

            # 如果区域发生变化或这是第一个点，创建新条目
            if current_entry is None or current_entry['region_key'] != region_key:
                # 保存前一个条目
                if current_entry is not None:
                    current_entry['end_time'] = point.time
                    timeline_entries.append({
                        'province': current_entry['province'],
                        'city': current_entry['city'],
                        'district': current_entry['district'],
                        'road_name': current_entry['road_name'],
                        'road_number': current_entry['road_number'],
                        'start_time': current_entry['start_time'],
                        'end_time': current_entry['end_time'],
                        'point_count': current_entry['point_count'],
                    })

                # 创建新条目
                current_entry = {
                    'region_key': region_key,
                    'province': point.province,
                    'city': point.city,
                    'district': point.district,
                    'road_name': point.road_name,
                    'road_number': point.road_number,
                    'start_time': point.time,
                    'end_time': None,
                    'point_count': 0,
                }
                point_count = 0

            current_entry['point_count'] = point_count + 1
            point_count += 1

        # 保存最后一个条目
        if current_entry is not None:
            current_entry['end_time'] = points[-1].time if points else None
            timeline_entries.append({
                'province': current_entry['province'],
                'city': current_entry['city'],
                'district': current_entry['district'],
                'road_name': current_entry['road_name'],
                'road_number': current_entry['road_number'],
                'start_time': current_entry['start_time'],
                'end_time': current_entry['end_time'],
                'point_count': current_entry['point_count'],
            })

        return timeline_entries

    def _calculate_distance(self, lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """计算两点之间的距离（米），使用 Haversine 公式"""
        from math import radians, sin, cos, sqrt, atan2

        dlat = radians(lat2 - lat1)
        dlon = radians(lon2 - lon1)
        a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))
        return 6371000 * c  # 地球半径 6371km

    def _aggregate_node_stats(self, node: dict) -> dict:
        """
        递归聚合节点统计信息，让上级包含下级的数据
        返回 (distance, point_count, start_time, end_time)
        """
        total_distance = node.get('own_distance', 0)
        total_points = node.get('own_point_count', 0)
        earliest_time = node.get('start_time')
        latest_time = node.get('end_time')
        earliest_index = node.get('start_index', -1)
        latest_index = node.get('end_index', -1)

        if node.get('children'):
            for child in node['children']:
                child_distance, child_points, child_start, child_end = self._aggregate_node_stats(child)
                total_distance += child_distance
                total_points += child_points

                # 聚合时间范围
                if child_start:
                    if earliest_time is None or child_start < earliest_time:
                        earliest_time = child_start
                if child_end:
                    if latest_time is None or child_end > latest_time:
                        latest_time = child_end

                # 聚合索引范围（父节点包含所有子节点的范围）
                child_start_index = child.get('start_index', -1)
                child_end_index = child.get('end_index', -1)
                if child_start_index >= 0:
                    if earliest_index < 0 or child_start_index < earliest_index:
                        earliest_index = child_start_index
                if child_end_index >= 0:
                    if latest_index < 0 or child_end_index > latest_index:
                        latest_index = child_end_index

        node['distance'] = total_distance
        node['point_count'] = total_points
        node['start_time'] = earliest_time
        node['end_time'] = latest_time
        # 更新父节点的索引范围（包含所有子节点）
        if node.get('children'):
            node['start_index'] = earliest_index
            node['end_index'] = latest_index

        return total_distance, total_points, earliest_time, latest_time

    async def get_region_tree(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
    ) -> dict:
        """
        获取轨迹的区域树结构（按时间顺序）

        返回按时间顺序展开的区域树，同一区域的多次经过会分开显示。
        结构：省 -> 市 -> 区 -> 道路
        每个节点包含统计信息（路径长度、时间范围）
        上级节点的统计数据包含所有下级节点

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID

        Returns:
            {'regions': 区域树列表, 'stats': 各级区域数量统计}
        """
        # 检查权限
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            return {'regions': [], 'stats': {'province': 0, 'city': 0, 'district': 0, 'road': 0}}

        # 获取轨迹点（按时间排序，实时记录场景下 point_index 可能乱序）
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.time, TrackPoint.created_at)
        )
        points = list(result.scalars().all())

        if not points:
            return {'regions': [], 'stats': {'province': 0, 'city': 0, 'district': 0, 'road': 0}}

        # 按时间顺序构建区域树
        root_nodes = []
        node_counter = [0]

        # 统计各级区域数量（去重）
        province_set = set()
        city_set = set()
        district_set = set()
        road_set = set()

        def create_node(name: str, node_type: str, road_number: str = None) -> dict:
            """创建一个新节点"""
            node_counter[0] += 1
            return {
                'id': f"node_{node_counter[0]}",
                'name': name,
                'type': node_type,
                'road_number': road_number,
                'own_distance': 0.0,  # 自己的路径长度（不含子节点）
                'distance': 0.0,  # 总路径长度（含子节点）
                'own_point_count': 0,  # 自己的点数
                'point_count': 0,  # 总点数
                'start_time': None,
                'end_time': None,
                'start_index': -1,  # 起始点索引（按时间顺序的位置，非 point_index）
                'end_index': -1,    # 结束点索引
                'children': [],
            }

        # 当前活跃的节点路径
        current_province = None
        current_city = None
        current_district = None
        current_road = None
        prev_point = None

        # 使用枚举索引作为时间顺序的位置（而非 point_index）
        # 实时记录场景下，point_index 不能保证时间顺序
        for time_idx, point in enumerate(points):
            province = point.province or '未知区域'
            city = point.city or province
            district = point.district or city
            road_name = point.road_name
            road_number = point.road_number

            # 统计各级区域（排除"未知区域"）
            if province and province != '未知区域': province_set.add(province)
            if city and city != province and city != '未知区域': city_set.add(city)
            if district and district != city and district != '未知区域': district_set.add(district)
            if road_name and road_name != '未知区域': road_set.add(road_name)

            # 检查是否需要创建新的省级节点
            if current_province is None or current_province[0] != province:
                # 先结束所有下层节点的索引范围
                if current_road is not None and prev_point is not None:
                    current_road[1]['end_index'] = time_idx - 1
                if current_district is not None and prev_point is not None:
                    current_district[1]['end_index'] = time_idx - 1
                if current_city is not None and prev_point is not None:
                    current_city[1]['end_index'] = time_idx - 1
                # 结束旧省级节点的索引范围
                if current_province is not None and prev_point is not None:
                    current_province[1]['end_index'] = time_idx - 1
                # 创建新省级节点并设置起始索引
                new_province = create_node(province, 'province')
                new_province['start_index'] = time_idx
                root_nodes.append(new_province)
                current_province = (province, new_province)
                current_city = None
                current_district = None
                current_road = None

            province_node = current_province[1]

            # 检查是否需要创建新的市级节点
            if current_city is None or current_city[0] != city:
                # 先结束所有下层节点的索引范围
                if current_road is not None and prev_point is not None:
                    current_road[1]['end_index'] = time_idx - 1
                if current_district is not None and prev_point is not None:
                    current_district[1]['end_index'] = time_idx - 1
                # 结束旧市级节点的索引范围
                if current_city is not None and prev_point is not None:
                    current_city[1]['end_index'] = time_idx - 1
                # 创建新市级节点并设置起始索引
                new_city = create_node(city, 'city')
                new_city['start_index'] = time_idx
                province_node['children'].append(new_city)
                current_city = (city, new_city)
                current_district = None
                current_road = None

            city_node = current_city[1] if current_city else province_node

            # 检查是否需要创建新的区级节点
            if current_district is None or current_district[0] != district:
                # 先结束所有下层节点的索引范围
                if current_road is not None and prev_point is not None:
                    current_road[1]['end_index'] = time_idx - 1
                # 结束旧区级节点的索引范围
                if current_district is not None and prev_point is not None:
                    current_district[1]['end_index'] = time_idx - 1
                # 创建新区级节点并设置起始索引
                new_district = create_node(district, 'district')
                new_district['start_index'] = time_idx
                city_node['children'].append(new_district)
                current_district = (district, new_district)
                current_road = None

            district_node = current_district[1] if current_district else city_node

            # 检查是否需要创建新的道路节点
            # 有道路信息时用道路名称，无道路信息时用"（无名）"
            if road_name:
                road_key = (road_name, road_number or '')
            else:
                road_key = ('（无名）', road_number or '')

            if current_road is None or current_road[0] != road_key:
                # 结束旧道路节点的索引范围
                if current_road is not None and prev_point is not None:
                    current_road[1]['end_index'] = time_idx - 1
                # 创建新道路节点并设置起始索引
                if road_name:
                    new_road = create_node(road_name, 'road', road_number)
                else:
                    new_road = create_node('（无名）', 'road', road_number)
                new_road['start_index'] = time_idx
                district_node['children'].append(new_road)
                current_road = (road_key, new_road)

            # 确定当前最低层级的活跃节点
            if current_road:
                active_node = current_road[1]
            elif current_district:
                active_node = current_district[1]
            elif current_city:
                active_node = current_city[1]
            else:
                active_node = province_node

            # 累加点数
            active_node['own_point_count'] += 1

            # 更新时间范围
            if point.time:
                if active_node['start_time'] is None or point.time < active_node['start_time']:
                    active_node['start_time'] = point.time
                if active_node['end_time'] is None or point.time > active_node['end_time']:
                    active_node['end_time'] = point.time

            # 计算距离
            if prev_point:
                distance = self._calculate_distance(
                    prev_point.latitude_wgs84, prev_point.longitude_wgs84,
                    point.latitude_wgs84, point.longitude_wgs84
                )
                active_node['own_distance'] += distance

            prev_point = point

        # 设置所有活跃节点的结束索引（使用最后的时间索引）
        last_time_idx = len(points) - 1
        if current_road is not None:
            current_road[1]['end_index'] = last_time_idx
        if current_district is not None:
            current_district[1]['end_index'] = last_time_idx
        if current_city is not None:
            current_city[1]['end_index'] = last_time_idx
        if current_province is not None:
            current_province[1]['end_index'] = last_time_idx

        # 后处理：聚合统计信息，让上级包含下级
        for node in root_nodes:
            self._aggregate_node_stats(node)

        return {
            'regions': root_nodes,
            'stats': {
                'province': len(province_set),
                'city': len(city_set),
                'district': len(district_set),
                'road': len(road_set),
            }
        }


    async def export_points_to_csv(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
    ) -> tuple[str, str]:
        """
        导出轨迹点为 CSV 格式（UTF-8 带 BOM）

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID

        Returns:
            (文件名, CSV 内容)
        """
        # 检查权限并获取轨迹
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            raise ValueError("轨迹不存在")

        # 获取轨迹点（按时间排序）
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.time, TrackPoint.created_at)
        )
        points = list(result.scalars().all())

        if not points:
            raise ValueError("轨迹没有数据点")

        # CSV BOM 头
        csv_lines = []
        csv_lines.append("\ufeff")  # UTF-8 BOM

        # CSV 表头
        headers = [
            "index", "time_date", "time_time", "time_microsecond", "elapsed_time",
            "longitude_wgs84", "latitude_wgs84",
            "longitude_gcj02", "latitude_gcj02",
            "longitude_bd09", "latitude_bd09",
            "elevation", "distance", "course", "speed",
            "province", "city", "area",
            "province_en", "city_en", "area_en",
            "road_num", "road_name", "road_name_en", "memo"
        ]
        csv_lines.append(",".join(headers))

        # 计算累计距离
        total_distance = 0.0
        prev_point = None

        for point in points:
            # 计算到前一个点的距离
            if prev_point:
                distance = self._calculate_distance(
                    prev_point.latitude_wgs84, prev_point.longitude_wgs84,
                    point.latitude_wgs84, point.longitude_wgs84
                )
                total_distance += distance
            else:
                total_distance = 0.0

            # 解析时间
            time_date = ""
            time_time = ""
            time_microsecond = ""
            elapsed_time = ""

            if point.time:
                time_date = point.time.strftime("%Y/%m/%d")
                time_time = point.time.strftime("%H:%M:%S")
                time_microsecond = str(point.time.microsecond)

                # 计算已用时间
                if track.start_time:
                    elapsed = (point.time - track.start_time).total_seconds()
                    elapsed_time = f"{elapsed:.2f}"

            # 构建行数据
            row = [
                str(point.point_index),
                time_date,
                time_time,
                time_microsecond,
                elapsed_time,
                f"{point.longitude_wgs84:.6f}" if point.longitude_wgs84 else "",
                f"{point.latitude_wgs84:.6f}" if point.latitude_wgs84 else "",
                f"{point.longitude_gcj02:.6f}" if point.longitude_gcj02 else "",
                f"{point.latitude_gcj02:.6f}" if point.latitude_gcj02 else "",
                f"{point.longitude_bd09:.6f}" if point.longitude_bd09 else "",
                f"{point.latitude_bd09:.6f}" if point.latitude_bd09 else "",
                f"{point.elevation:.1f}" if point.elevation is not None else "",
                f"{total_distance:.2f}",
                f"{point.bearing:.2f}" if point.bearing is not None else "",
                f"{point.speed:.2f}" if point.speed is not None else "",
                point.province or "",
                point.city or "",
                point.district or "",
                point.province_en or "",
                point.city_en or "",
                point.district_en or "",
                point.road_number or "",
                point.road_name or "",
                point.road_name_en or "",
                getattr(point, 'memo', None) or "",
            ]

            # CSV 转义：包含逗号的字段用引号包裹
            csv_row = []
            for field in row:
                if "," in field or '"' in field or "\n" in field:
                    # 转义引号并包裹
                    field = '"' + field.replace('"', '""') + '"'
                csv_row.append(field)

            csv_lines.append(",".join(csv_row))
            prev_point = point

        filename = f"{track.name}_points.csv"
        content = "\n".join(csv_lines)

        return filename, content

    async def export_points_to_xlsx(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
    ) -> tuple[str, bytes]:
        """
        导出轨迹点为 XLSX 格式

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID

        Returns:
            (文件名, XLSX 二进制内容)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO

        # 检查权限并获取轨迹
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            raise ValueError("轨迹不存在")

        # 获取轨迹点（按时间排序）
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.time, TrackPoint.created_at)
        )
        points = list(result.scalars().all())

        if not points:
            raise ValueError("轨迹没有数据点")

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "轨迹点"

        # 设置表头
        headers = [
            "index", "time_date", "time_time", "time_microsecond", "elapsed_time",
            "longitude_wgs84", "latitude_wgs84",
            "longitude_gcj02", "latitude_gcj02",
            "longitude_bd09", "latitude_bd09",
            "elevation", "distance", "course", "speed",
            "province", "city", "area",
            "province_en", "city_en", "area_en",
            "road_num", "road_name", "road_name_en", "memo"
        ]

        # 写入表头（加粗）
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 计算累计距离并写入数据
        total_distance = 0.0
        prev_point = None

        for row_idx, point in enumerate(points, start=2):
            # 计算到前一个点的距离
            if prev_point:
                distance = self._calculate_distance(
                    prev_point.latitude_wgs84, prev_point.longitude_wgs84,
                    point.latitude_wgs84, point.longitude_wgs84
                )
                total_distance += distance
            else:
                total_distance = 0.0

            # 解析时间
            time_date = ""
            time_time = ""
            time_microsecond = ""
            elapsed_time = ""

            if point.time:
                time_date = point.time.strftime("%Y/%m/%d")
                time_time = point.time.strftime("%H:%M:%S")
                time_microsecond = str(point.time.microsecond)

                # 计算已用时间
                if track.start_time:
                    elapsed = (point.time - track.start_time).total_seconds()
                    elapsed_time = f"{elapsed:.2f}"

            # 写入行数据
            row_data = [
                point.point_index,
                time_date,
                time_time,
                time_microsecond,
                elapsed_time,
                round(point.longitude_wgs84, 6) if point.longitude_wgs84 else None,
                round(point.latitude_wgs84, 6) if point.latitude_wgs84 else None,
                round(point.longitude_gcj02, 6) if point.longitude_gcj02 else None,
                round(point.latitude_gcj02, 6) if point.latitude_gcj02 else None,
                round(point.longitude_bd09, 6) if point.longitude_bd09 else None,
                round(point.latitude_bd09, 6) if point.latitude_bd09 else None,
                round(point.elevation, 1) if point.elevation is not None else None,
                round(total_distance, 2),
                round(point.bearing, 2) if point.bearing is not None else None,
                round(point.speed, 2) if point.speed is not None else None,
                point.province,
                point.city,
                point.district,
                point.province_en,
                point.city_en,
                point.district_en,
                point.road_number,
                point.road_name,
                point.road_name_en,
                getattr(point, 'memo', None),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

            prev_point = point

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{track.name}_points.xlsx"
        return filename, output.read()

    async def export_points_to_kml(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
        crs: Optional[str] = None,
    ) -> tuple[str, str]:
        """
        导出轨迹点为 KML 格式（两步路兼容格式）

        使用 Google gx:Track 扩展格式，支持时间信息

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID
            crs: 坐标系 (original/wgs84/gcj02/bd09)，默认为 original

        Returns:
            (文件名, KML 内容)
        """
        # 检查权限并获取轨迹
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            raise ValueError("轨迹不存在")

        # 获取轨迹点（按时间排序）
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.time, TrackPoint.created_at)
        )
        points = list(result.scalars().all())

        if not points:
            raise ValueError("轨迹没有数据点")

        # 确定使用的坐标系
        if crs == "original" or crs is None:
            original_crs_value = track.original_crs or "wgs84"
            target_crs = str(original_crs_value)
        else:
            target_crs = crs

        # 构建 KML 内容
        kml_lines = []

        # KML 头部
        kml_lines.append('<?xml version="1.0" encoding="UTF-8"?>')
        kml_lines.append('<kml xmlns="http://www.opengis.net/kml/2.2" xmlns:gx="http://www.google.com/kml/ext/2.2">')
        kml_lines.append('  <Document>')
        kml_lines.append(f'    <name>{track.name}</name>')
        if track.description:
            kml_lines.append(f'    <description>{track.description}</description>')

        # 添加样式（与两步路类似）
        kml_lines.append('    <Style id="TrackStyle_n">')
        kml_lines.append('      <LineStyle>')
        kml_lines.append('        <color>ffff0000</color>')
        kml_lines.append('        <width>5</width>')
        kml_lines.append('      </LineStyle>')
        kml_lines.append('    </Style>')

        kml_lines.append('    <Style id="TrackStyle_h">')
        kml_lines.append('      <LineStyle>')
        kml_lines.append('        <color>ff0000ff</color>')
        kml_lines.append('        <width>7</width>')
        kml_lines.append('      </LineStyle>')
        kml_lines.append('    </Style>')

        kml_lines.append('    <StyleMap id="TrackStyle">')
        kml_lines.append('      <Pair>')
        kml_lines.append('        <key>normal</key>')
        kml_lines.append('        <styleUrl>#TrackStyle_n</styleUrl>')
        kml_lines.append('      </Pair>')
        kml_lines.append('      <Pair>')
        kml_lines.append('        <key>highlight</key>')
        kml_lines.append('        <styleUrl>#TrackStyle_h</styleUrl>')
        kml_lines.append('      </Pair>')
        kml_lines.append('    </StyleMap>')

        # 开始 Placemark
        kml_lines.append('    <Placemark>')
        kml_lines.append(f'      <name>{track.name}</name>')
        kml_lines.append('      <styleUrl>#TrackStyle</styleUrl>')
        kml_lines.append('      <gx:Track>')

        # 添加坐标（根据选择的坐标系）
        for point in points:
            if target_crs == "wgs84":
                lon, lat = point.longitude_wgs84, point.latitude_wgs84
            elif target_crs == "gcj02":
                lon, lat = point.longitude_gcj02, point.latitude_gcj02
            elif target_crs == "bd09":
                lon, lat = point.longitude_bd09, point.latitude_bd09
            else:
                # 默认使用 WGS84
                lon, lat = point.longitude_wgs84, point.latitude_wgs84
            elev = point.elevation if point.elevation is not None else 0
            kml_lines.append(f'        <gx:coord>{lon} {lat} {elev}</gx:coord>')

        # 添加时间
        for point in points:
            if point.time:
                # ISO 8601 格式，带 Z 后缀表示 UTC
                time_str = point.time.strftime('%Y-%m-%dT%H:%M:%SZ')
                kml_lines.append(f'        <when>{time_str}</when>')
            else:
                # 如果没有时间，使用占位符（两步路可能不支持无时间的点）
                kml_lines.append('        <when></when>')

        kml_lines.append('      </gx:Track>')
        kml_lines.append('    </Placemark>')

        # KML 尾部
        kml_lines.append('  </Document>')
        kml_lines.append('</kml>')

        # 生成文件名（使用轨迹名称）
        # 如果坐标系不是原坐标系，在文件名中加上坐标系后缀
        original_crs_str = str(track.original_crs or "wgs84")
        target_crs_str = str(target_crs) if target_crs else "wgs84"
        crs_suffix = "" if target_crs_str == original_crs_str else f"_{target_crs_str.upper()}"
        filename = f"{track.name}{crs_suffix}.kml"
        content = '\n'.join(kml_lines)

        return filename, content

    async def import_points_from_file(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
        file_content: bytes,
        file_format: str,
        match_mode: str = "index",
        timezone: str = "UTC",
        time_tolerance: float = 1.0,
    ) -> dict:
        """
        从文件导入轨迹点数据，更新行政区划和道路信息

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID
            file_content: 文件内容
            file_format: 文件格式 (csv 或 xlsx)
            match_mode: 匹配方式 (index=索引匹配, time=时间匹配)
            timezone: 导入文件的时间戳时区（如 UTC、UTC+8、Asia/Shanghai）
            time_tolerance: 时间匹配误差（秒），默认 1 秒（不含）

        Returns:
            {"updated": 更新的点数, "total": 总点数, "matched_by": 匹配方式}
        """
        # 检查权限
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            raise ValueError("轨迹不存在")

        # 获取现有轨迹点（按时间排序）
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.time, TrackPoint.created_at)
        )
        points = list(result.scalars().all())

        if not points:
            raise ValueError("轨迹没有数据点")

        # 创建索引到点的映射
        # 注意：在 index 模式下，使用时间顺序的位置（0, 1, 2...）而非 point_index
        # 因为实时记录场景下 point_index 不能保证时间顺序
        points_map = {idx: p for idx, p in enumerate(points)}

        # 解析时区
        from datetime import datetime, timezone as dt_timezone, timedelta
        import zoneinfo

        # 解析时区字符串
        try:
            if timezone.startswith("UTC") or timezone.startswith("GMT"):
                # 处理 UTC+8, UTC-5 等格式
                sign = 1
                offset_str = timezone[3:].strip()
                if offset_str:
                    if offset_str[0] == "-":
                        sign = -1
                        offset_str = offset_str[1:]
                    hours = int(offset_str)
                    tz = dt_timezone(timedelta(hours=sign * hours))
                else:
                    tz = dt_timezone.utc
            else:
                # 使用 IANA 时区数据库（如 Asia/Shanghai）
                tz = zoneinfo.ZoneInfo(timezone)
        except Exception as e:
            logger.warning(f"Invalid timezone '{timezone}', using UTC: {e}")
            tz = dt_timezone.utc

        # 创建时间到点的映射（用于时间匹配）
        # 使用 (time.timestamp(), 误差范围内) 的方式存储
        points_by_time = {}
        db_time_range = {"min": None, "max": None}
        for p in points:
            if p.time:
                # 数据库中的 time 字段是 UTC 时间（无时区信息）
                # 需要明确指定为 UTC 时区再获取时间戳，否则会被当作本地时间处理
                if p.time.tzinfo is None:
                    ts = p.time.replace(tzinfo=dt_timezone.utc).timestamp()
                else:
                    ts = p.time.timestamp()
                points_by_time[ts] = p
                # 记录时间范围
                if db_time_range["min"] is None or ts < db_time_range["min"]:
                    db_time_range["min"] = ts
                if db_time_range["max"] is None or ts > db_time_range["max"]:
                    db_time_range["max"] = ts

        # 调试：显示数据库中的前 3 个时间戳
        sample_timestamps = list(points_by_time.keys())[:3]
        logger.info(f"Database sample timestamps: {sample_timestamps}")

        # 记录使用的匹配方式
        matched_by = "none"

        # 调试日志 - 显示数据库时间范围
        if db_time_range["min"] is not None:
            from datetime import datetime as dt
            min_time = dt.fromtimestamp(db_time_range["min"], dt_timezone.utc)
            max_time = dt.fromtimestamp(db_time_range["max"], dt_timezone.utc)
            logger.info(f"Import: match_mode={match_mode}, timezone={timezone}, total_points={len(points)}")
            logger.info(f"Database time range (UTC): {min_time} to {max_time}")
        else:
            logger.info(f"Import: match_mode={match_mode}, timezone={timezone}, total_points={len(points)}")
            logger.warning("No time data found in database points!")

        # 计数器用于限制日志输出
        row_count = [0]  # 用列表以便在嵌套函数中修改

        def find_point_by_time(parsed_time: datetime) -> TrackPoint | None:
            """根据时间查找匹配的轨迹点"""
            if not parsed_time:
                return None

            # 将文件时间（无时区）解释为用户指定的时区，然后转换为 UTC 时间戳
            # 文件中的时间是 UTC+8 时间，需要转换为 UTC 来匹配数据库
            # 例如：文件时间 15:21:47 (UTC+8) -> 07:21:47 UTC
            if parsed_time.tzinfo is None:
                # 文件时间是用户指定的时区（如 UTC+8）
                parsed_time = parsed_time.replace(tzinfo=tz)

            # 转换为 UTC 时间戳
            target_ts = parsed_time.astimezone(dt_timezone.utc).timestamp()

            # 只记录前 3 行的时间匹配日志
            if row_count[0] < 3:
                logger.info(f"find_point_by_time: file_time={parsed_time} -> target_ts={target_ts}")

            # 查找最接近的点（使用用户指定的误差范围，不含边界值）
            for ts, p in points_by_time.items():
                diff = abs(ts - target_ts)
                if diff < time_tolerance:
                    if row_count[0] < 3:
                        logger.info(f"  Found match: ts={ts}, diff={diff:.3f}s")
                    return p

            return None

        def parse_time_from_row(row: dict, headers: set | list | None = None) -> datetime | None:
            """从行数据解析时间"""
            # 优先使用 time_date + time_time 组合
            time_date = None
            time_time = None
            time_col = None

            if isinstance(row, dict):
                # 记录第一行用于调试
                if row.get("index") == "0" or row.get("index") == 0:
                    logger.info(f"CSV first row keys: {list(row.keys())}")
                    logger.info(f"CSV first row: {row}")
                time_date = row.get("time_date", "").strip()
                time_time = row.get("time_time", "").strip()
                time_col = row.get("time", "").strip()
            else:
                # XLSX 使用 headers
                if headers:
                    # 如果 headers 已经是 list，直接使用；否则转换
                    header_list = headers if isinstance(headers, list) else list(headers)
                    # 记录第一行用于调试
                    if "index" in header_list:
                        idx = header_list.index("index")
                        if idx < len(row) and (row[idx] == "0" or row[idx] == 0):
                            logger.info(f"XLSX headers: {header_list}")
                            logger.info(f"XLSX first row: {dict(zip(header_list, row))}")

                    if "time_date" in header_list:
                        idx = header_list.index("time_date")
                        if idx < len(row):
                            time_date = str(row[idx]).strip() if row[idx] else ""
                    if "time_time" in header_list:
                        idx = header_list.index("time_time")
                        if idx < len(row):
                            time_time = str(row[idx]).strip() if row[idx] else ""
                    if "time" in header_list:
                        idx = header_list.index("time")
                        if idx < len(row) and row[idx] is not None:
                            # XLSX 中可能是 datetime 对象，直接返回
                            if isinstance(row[idx], datetime):
                                return row[idx]
                            time_col = str(row[idx]).strip()

            # 尝试 time_date + time_time 组合
            if time_date and time_time:
                try:
                    dt_str = f"{time_date} {time_time}"
                    # 支持多种日期格式
                    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                                "%Y/%m/%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S.%f"):
                        try:
                            return datetime.strptime(dt_str, fmt)
                        except ValueError:
                            continue
                except Exception:
                    pass

            # 尝试单一 time 列
            if time_col:
                # 只记录前 3 行的解析日志
                if row_count[0] < 3:
                    logger.info(f"Parsing time_col: '{time_col}'")
                # 尝试常见格式
                for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S",
                            "%Y/%m/%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S.%f",
                            "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f",
                            "%Y/%d/%m %H:%M:%S", "%Y/%d/%m %H:%M:%S.%f"):
                    try:
                        result = datetime.strptime(time_col, fmt)
                        if row_count[0] < 3:
                            logger.info(f"Successfully parsed with fmt '{fmt}': {result}")
                        return result
                    except ValueError:
                        continue
                if row_count[0] < 3:
                    logger.warning(f"Failed to parse time: '{time_col}'")

            return None

        def update_point_fields(point: TrackPoint, row: dict, headers: set | list | None = None):
            """更新点的可编辑字段"""
            def get_val(key: str) -> str | None:
                if isinstance(row, dict):
                    val = row.get(key)
                    return val.strip() if val else None
                else:
                    if headers and key in headers:
                        # 如果 headers 已经是 list，直接使用；否则转换
                        header_list = headers if isinstance(headers, list) else list(headers)
                        idx = header_list.index(key)
                        if idx < len(row):
                            val = row[idx]
                            return str(val).strip() if val else None
                return None

            if get_val("province") is not None:
                point.province = get_val("province") or None
            if get_val("city") is not None:
                point.city = get_val("city") or None
            if get_val("area") is not None:
                point.district = get_val("area") or None
            if get_val("province_en") is not None:
                point.province_en = get_val("province_en") or None
            if get_val("city_en") is not None:
                point.city_en = get_val("city_en") or None
            if get_val("area_en") is not None:
                point.district_en = get_val("area_en") or None
            if get_val("road_num") is not None:
                point.road_number = get_val("road_num") or None
            if get_val("road_name") is not None:
                point.road_name = get_val("road_name") or None
            if get_val("road_name_en") is not None:
                point.road_name_en = get_val("road_name_en") or None
            if get_val("memo") is not None:
                point.memo = get_val("memo") or None

            point.updated_by = user_id

        # 解析文件
        if file_format == "csv":
            import csv
            from io import StringIO

            # 处理 BOM
            content_str = file_content.decode('utf-8-sig')
            reader = csv.DictReader(StringIO(content_str))

            updated_count = 0
            for row in reader:
                point = None

                # 根据匹配方式选择匹配逻辑
                if match_mode == "index":
                    # 使用 index 匹配
                    index_str = row.get("index", "").strip()
                    if index_str:
                        try:
                            index = int(index_str)
                            if index in points_map:
                                point = points_map[index]
                                matched_by = "index"
                        except ValueError:
                            pass
                else:  # match_mode == "time"
                    # 使用时间匹配
                    parsed_time = parse_time_from_row(row)
                    row_count[0] += 1  # 增加计数
                    if parsed_time:
                        point = find_point_by_time(parsed_time)
                        if point:
                            matched_by = "time"

                if point:
                    update_point_fields(point, row)
                    updated_count += 1

        elif file_format == "xlsx":
            from openpyxl import load_workbook
            from io import BytesIO

            wb = load_workbook(BytesIO(file_content))
            ws = wb.active

            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            headers_set = set(headers)
            logger.info(f"XLSX file headers: {headers}")

            updated_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(v is None for v in row):
                    continue

                point = None

                # 根据匹配方式选择匹配逻辑
                if match_mode == "index":
                    # 使用 index 匹配
                    if "index" in headers_set:
                        idx = headers.index("index")
                        if idx < len(row) and row[idx] is not None:
                            try:
                                index = int(row[idx])
                                if index in points_map:
                                    point = points_map[index]
                                    matched_by = "index"
                            except (ValueError, TypeError):
                                pass
                else:  # match_mode == "time"
                    # 使用时间匹配
                    parsed_time = parse_time_from_row(row, headers)
                    row_count[0] += 1  # 增加计数
                    if parsed_time:
                        point = find_point_by_time(parsed_time)
                        if point:
                            matched_by = "time"

                if point:
                    update_point_fields(point, row, headers)
                    updated_count += 1
        else:
            raise ValueError(f"不支持的文件格式: {file_format}")

        # 检查是否有行政区划或道路信息，更新 track 状态
        has_area = any(
            p.province or p.city or p.district or
            p.province_en or p.city_en or p.district_en
            for p in points
        )
        has_road = any(
            p.road_number or p.road_name or p.road_name_en
            for p in points
        )

        if has_area:
            track.has_area_info = True
        if has_road:
            track.has_road_info = True
        track.updated_by = user_id

        await db.commit()

        logger.info(f"Imported data for track {track_id}: {updated_count}/{len(points)} points updated, matched by {matched_by}")

        return {
            "updated": updated_count,
            "total": len(points),
            "matched_by": matched_by
        }


    async def create_from_csv(
        self,
        db: AsyncSession,
        user: User,
        filename: str,
        csv_content: str,
        name: str,
        description: Optional[str] = None,
        original_crs: CoordinateType = 'wgs84',
        convert_to: Optional[CoordinateType] = None,
        fill_geocoding: bool = False,
    ) -> Track:
        """
        从 CSV 内容创建轨迹

        支持两种 CSV 格式：

        1. GPS Logger 应用格式：
           - time: ISO 8601 格式时间（UTC），如 2026-01-29T14:20:28.000Z
           - lat: 纬度
           - lon: 经度
           - elevation: 海拔（米）
           - speed: 速度（m/s）
           - bearing: 方向角（度）

        2. 本项目导出格式：
           - index, time_date, time_time, time_microsecond, elapsed_time
           - longitude_wgs84, latitude_wgs84
           - longitude_gcj02, latitude_gcj02
           - longitude_bd09, latitude_bd09
           - elevation, distance, course, speed
           - province, city, area, province_en, city_en, area_en
           - road_num, road_name, road_name_en, memo

        Args:
            db: 数据库会话
            user: 用户对象
            filename: 原始文件名
            csv_content: CSV 文件内容
            name: 轨迹名称
            description: 轨迹描述
            original_crs: 原始坐标系（仅用于 GPS Logger 格式）
            convert_to: 转换到目标坐标系（仅用于 GPS Logger 格式）
            fill_geocoding: 是否填充行政区划和道路信息（仅用于 GPS Logger 格式）

        Returns:
            创建的轨迹对象
        """
        import csv
        from io import StringIO
        from datetime import datetime

        # 解析 CSV
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        if not rows:
            raise ValueError("CSV 文件中没有数据")

        # 检测 CSV 格式类型
        is_project_export = (
            reader.fieldnames and
            'longitude_wgs84' in reader.fieldnames and
            'latitude_wgs84' in reader.fieldnames
        )

        if is_project_export:
            # 使用本项目导出格式解析
            return await self._create_from_csv_project_format(
                db, user, filename, rows, name, description
            )

        # GPS Logger 格式
        # 验证必需的列
        required_columns = ['time', 'lat', 'lon']
        for col in required_columns:
            if col not in reader.fieldnames:
                raise ValueError(f"CSV 文件缺少必需的列: {col}")

        # 计算统计信息
        total_distance = 0
        elevation_gain = 0
        elevation_loss = 0
        prev_elevation = None
        start_time = None
        end_time = None

        points_data = []
        prev_point_data = None

        for idx, row in enumerate(rows):
            # 解析时间
            point_time = None
            time_str = row.get('time', '').strip()
            if time_str:
                try:
                    # GPS Logger 使用 ISO 8601 格式，如 2026-01-29T14:20:28.000Z
                    # 解析为 UTC 时间
                    point_time = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
                    # 移除时区信息（与数据库保持一致）
                    point_time = point_time.replace(tzinfo=None)

                    if start_time is None:
                        start_time = point_time
                    end_time = point_time
                except ValueError as e:
                    logger.warning(f"无法解析时间 '{time_str}': {e}")

            # 解析坐标
            try:
                lat = float(row['lat'])
                lon = float(row['lon'])
            except (ValueError, KeyError) as e:
                logger.warning(f"跳过无效坐标的行 {idx}: {e}")
                continue

            # 坐标转换
            all_coords = convert_point_to_all(lon, lat, original_crs)

            # 解析海拔
            elevation = None
            elev_str = row.get('elevation', '').strip()
            if elev_str:
                try:
                    elevation = float(elev_str)
                    # 海拔变化
                    if prev_elevation is not None:
                        diff = elevation - prev_elevation
                        if diff > 0:
                            elevation_gain += diff
                        else:
                            elevation_loss += abs(diff)
                    prev_elevation = elevation
                except ValueError:
                    pass

            # 解析速度（CSV 中已有）
            speed = None
            speed_str = row.get('speed', '').strip()
            if speed_str:
                try:
                    speed = float(speed_str)
                except ValueError:
                    pass

            # 解析方向角（CSV 中已有）
            bearing = None
            bearing_str = row.get('bearing', '').strip()
            if bearing_str:
                try:
                    bearing = float(bearing_str)
                except ValueError:
                    pass

            # 如果 CSV 没有提供速度或方向角，计算它们
            distance_from_prev = 0
            if idx > 0 and prev_point_data:
                # 使用 WGS84 坐标计算 3D 距离
                from math import sqrt, radians, sin, cos, atan2

                lat1 = radians(prev_point_data['latitude_wgs84'])
                lon1 = radians(prev_point_data['longitude_wgs84'])
                lat2 = radians(all_coords['wgs84'][1])
                lon2 = radians(all_coords['wgs84'][0])

                # Haversine 公式计算水平距离
                dlat = lat2 - lat1
                dlon = lon2 - lon1
                a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
                c = 2 * atan2(sqrt(a), sqrt(1 - a))
                horizontal_distance = 6371000 * c  # 地球半径 6371km

                # 计算垂直距离
                elev1 = prev_point_data.get('elevation') or 0
                elev2 = elevation or 0
                vertical_distance = elev2 - elev1

                # 3D 距离
                distance_from_prev = sqrt(horizontal_distance ** 2 + vertical_distance ** 2)
                total_distance += distance_from_prev

                # 如果 CSV 没有提供速度，计算速度
                if speed is None:
                    speed = self._calculate_speed(point_time, prev_point_data['time'], distance_from_prev)

                # 如果 CSV 没有提供方向角，计算方向角
                if bearing is None:
                    bearing = self._calculate_bearing(
                        prev_point_data['latitude_wgs84'],
                        prev_point_data['longitude_wgs84'],
                        all_coords['wgs84'][1],
                        all_coords['wgs84'][0]
                    )

            point_data = {
                'point_index': idx,
                'time': point_time,
                'latitude_wgs84': all_coords['wgs84'][1],
                'longitude_wgs84': all_coords['wgs84'][0],
                'latitude_gcj02': all_coords['gcj02'][1],
                'longitude_gcj02': all_coords['gcj02'][0],
                'latitude_bd09': all_coords['bd09'][1],
                'longitude_bd09': all_coords['bd09'][0],
                'elevation': elevation,
                'speed': speed,
                'bearing': bearing if idx > 0 else None,  # 第一个点没有方位角
            }
            points_data.append(point_data)
            prev_point_data = point_data

        if not points_data:
            raise ValueError("CSV 文件中没有有效的轨迹点")

        # 计算时长（秒）
        duration = 0
        if start_time and end_time:
            duration = int((end_time - start_time).total_seconds())

        # 创建轨迹记录，带审计字段
        track_obj = Track(
            user_id=user.id,
            name=name,
            description=description,
            original_filename=filename,
            original_crs=original_crs,
            distance=round(total_distance, 2),
            duration=duration,
            elevation_gain=round(elevation_gain, 2),
            elevation_loss=round(elevation_loss, 2),
            start_time=start_time,
            end_time=end_time,
            has_area_info=False,
            has_road_info=False,
            created_by=user.id,
            updated_by=user.id,
            is_valid=True,
        )
        db.add(track_obj)
        await db.flush()  # 获取 track_id

        # 批量创建轨迹点
        insert_values = []
        for point_data in points_data:
            insert_values.append({
                "track_id": track_obj.id,
                "point_index": point_data["point_index"],
                "time": point_data["time"],
                "latitude_wgs84": point_data["latitude_wgs84"],
                "longitude_wgs84": point_data["longitude_wgs84"],
                "latitude_gcj02": point_data["latitude_gcj02"],
                "longitude_gcj02": point_data["longitude_gcj02"],
                "latitude_bd09": point_data["latitude_bd09"],
                "longitude_bd09": point_data["longitude_bd09"],
                "elevation": point_data["elevation"],
                "speed": point_data["speed"],
                "bearing": point_data["bearing"],
                "province": None,
                "city": None,
                "district": None,
                "province_en": None,
                "city_en": None,
                "district_en": None,
                "road_name": None,
                "road_number": None,
                "road_name_en": None,
                "created_by": user.id,
                "updated_by": user.id,
                "is_valid": True,
            })

        # 分批插入，每批 500 条
        batch_size = 500
        for i in range(0, len(insert_values), batch_size):
            batch = insert_values[i:i + batch_size]
            try:
                await db.execute(
                    TrackPoint.__table__.insert().values(batch)
                )
            except Exception as e:
                logger.error(f"Error inserting batch {i//batch_size}: {e}")
                raise

        await db.commit()
        await db.refresh(track_obj)

        # 异步填充地理信息（如果启用）
        if fill_geocoding:
            task = asyncio.create_task(
                self.fill_geocoding_info(db, track_obj.id, user.id)
            )
            task.add_done_callback(self._background_tasks.discard)
            self._background_tasks.add(task)

        return track_obj

    async def _create_from_csv_project_format(
        self,
        db: AsyncSession,
        user: User,
        filename: str,
        rows: list,
        name: str,
        description: Optional[str] = None,
    ) -> Track:
        """
        从本项目导出的 CSV 格式创建轨迹

        支持的字段：
        - index, time_date, time_time, time_microsecond, elapsed_time
        - longitude_wgs84, latitude_wgs84
        - longitude_gcj02, latitude_gcj02
        - longitude_bd09, latitude_bd09
        - elevation, distance, course, speed
        - province, city, area, province_en, city_en, area_en
        - road_num, road_name, road_name_en, memo
        """
        from datetime import datetime

        # 计算统计信息
        total_distance = 0
        elevation_gain = 0
        elevation_loss = 0
        prev_elevation = None
        start_time = None
        end_time = None

        points_data = []
        prev_point_data = None

        # 用于检测是否有地理信息
        has_area_info = False
        has_road_info = False

        for idx, row in enumerate(rows):
            # 解析时间
            point_time = None
            time_date = row.get('time_date', '').strip()
            time_time = row.get('time_time', '').strip()

            if time_date and time_time:
                try:
                    # 格式: 2026/01/29 和 14:20:28
                    time_str = f"{time_date} {time_time}"
                    point_time = datetime.strptime(time_str, "%Y/%m/%d %H:%M:%S")

                    # 解析微秒
                    time_microsecond = row.get('time_microsecond', '').strip()
                    if time_microsecond and time_microsecond.isdigit():
                        point_time = point_time.replace(microsecond=int(time_microsecond))

                    if start_time is None:
                        start_time = point_time
                    end_time = point_time
                except ValueError as e:
                    logger.warning(f"无法解析时间 '{time_date} {time_time}': {e}")

            # 解析多坐标系坐标
            try:
                lon_wgs84 = self._parse_float(row.get('longitude_wgs84'))
                lat_wgs84 = self._parse_float(row.get('latitude_wgs84'))
                lon_gcj02 = self._parse_float(row.get('longitude_gcj02'))
                lat_gcj02 = self._parse_float(row.get('latitude_gcj02'))
                lon_bd09 = self._parse_float(row.get('longitude_bd09'))
                lat_bd09 = self._parse_float(row.get('latitude_bd09'))

                # 如果 WGS84 坐标缺失，尝试从其他坐标系转换
                if lon_wgs84 is None or lat_wgs84 is None:
                    if lon_gcj02 is not None and lat_gcj02 is not None:
                        # 从 GCJ02 转换到 WGS84
                        from app.gpxutil_wrapper.coord_transform import gcj02_to_wgs84
                        lon_wgs84, lat_wgs84 = gcj02_to_wgs84(lon_gcj02, lat_gcj02)
                    elif lon_bd09 is not None and lat_bd09 is not None:
                        # 从 BD09 转换到 WGS84
                        from app.gpxutil_wrapper.coord_transform import bd09_to_wgs84
                        lon_wgs84, lat_wgs84 = bd09_to_wgs84(lon_bd09, lat_bd09)
                    else:
                        raise ValueError("缺少有效坐标")

                # 如果其他坐标系缺失，从 WGS84 转换
                if lon_gcj02 is None or lat_gcj02 is None:
                    from app.gpxutil_wrapper.coord_transform import wgs84_to_gcj02
                    lon_gcj02, lat_gcj02 = wgs84_to_gcj02(lon_wgs84, lat_wgs84)

                if lon_bd09 is None or lat_bd09 is None:
                    from app.gpxutil_wrapper.coord_transform import wgs84_to_bd09
                    lon_bd09, lat_bd09 = wgs84_to_bd09(lon_wgs84, lat_wgs84)

            except (ValueError, TypeError) as e:
                logger.warning(f"跳过无效坐标的行 {idx}: {e}")
                continue

            # 解析海拔
            elevation = self._parse_float(row.get('elevation'))

            # 计算海拔变化
            if elevation is not None and prev_elevation is not None:
                diff = elevation - prev_elevation
                if diff > 0:
                    elevation_gain += diff
                else:
                    elevation_loss += abs(diff)
            prev_elevation = elevation

            # 解析速度和方向角（从 CSV 中读取）
            speed = self._parse_float(row.get('speed'))
            bearing = self._parse_float(row.get('course'))  # course 即 bearing

            # 解析地理信息
            province = row.get('province', '').strip() or None
            city = row.get('city', '').strip() or None
            district = row.get('area', '').strip() or None  # area 对应 district
            province_en = row.get('province_en', '').strip() or None
            city_en = row.get('city_en', '').strip() or None
            district_en = row.get('area_en', '').strip() or None
            road_number = row.get('road_num', '').strip() or None
            road_name = row.get('road_name', '').strip() or None
            road_name_en = row.get('road_name_en', '').strip() or None

            # 检测是否有地理信息
            if province or city or district or province_en or city_en or district_en:
                has_area_info = True
            if road_number or road_name or road_name_en:
                has_road_info = True

            # 计算距离和速度（如果未提供）
            distance_from_prev = 0
            if idx > 0 and prev_point_data:
                # 使用 WGS84 坐标计算 3D 距离
                from math import sqrt, radians, sin, cos, atan2

                lat1 = radians(prev_point_data['latitude_wgs84'])
                lon1 = radians(prev_point_data['longitude_wgs84'])
                lat2 = radians(lat_wgs84)
                lon2 = radians(lon_wgs84)

                # Haversine 公式计算水平距离
                dlat = lat2 - lat1
                dlon = lon2 - lon1
                a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
                c = 2 * atan2(sqrt(a), sqrt(1 - a))
                horizontal_distance = 6371000 * c  # 地球半径 6371km

                # 计算垂直距离
                elev1 = prev_point_data.get('elevation') or 0
                elev2 = elevation or 0
                vertical_distance = elev2 - elev1

                # 3D 距离
                distance_from_prev = sqrt(horizontal_distance ** 2 + vertical_distance ** 2)
                total_distance += distance_from_prev

                # 如果 CSV 没有提供速度，计算速度
                if speed is None:
                    speed = self._calculate_speed(point_time, prev_point_data['time'], distance_from_prev)

                # 如果 CSV 没有提供方向角，计算方向角
                if bearing is None:
                    bearing = self._calculate_bearing(
                        prev_point_data['latitude_wgs84'],
                        prev_point_data['longitude_wgs84'],
                        lat_wgs84,
                        lon_wgs84
                    )

            # 获取索引
            point_index = idx
            if 'index' in row and row['index']:
                try:
                    point_index = int(row['index'])
                except ValueError:
                    pass

            point_data = {
                'point_index': point_index,
                'time': point_time,
                'latitude_wgs84': lat_wgs84,
                'longitude_wgs84': lon_wgs84,
                'latitude_gcj02': lat_gcj02,
                'longitude_gcj02': lon_gcj02,
                'latitude_bd09': lat_bd09,
                'longitude_bd09': lon_bd09,
                'elevation': elevation,
                'speed': speed,
                'bearing': bearing if idx > 0 else None,  # 第一个点没有方位角
                'province': province,
                'city': city,
                'district': district,
                'province_en': province_en,
                'city_en': city_en,
                'district_en': district_en,
                'road_number': road_number,
                'road_name': road_name,
                'road_name_en': road_name_en,
            }
            points_data.append(point_data)
            prev_point_data = point_data

        if not points_data:
            raise ValueError("CSV 文件中没有有效的轨迹点")

        # 计算时长（秒）
        duration = 0
        if start_time and end_time:
            duration = int((end_time - start_time).total_seconds())

        # 确定原始坐标系（使用 WGS84 作为默认，因为导出包含所有坐标系）
        original_crs = 'wgs84'

        # 创建轨迹记录，带审计字段
        track_obj = Track(
            user_id=user.id,
            name=name,
            description=description,
            original_filename=filename,
            original_crs=original_crs,
            distance=round(total_distance, 2),
            duration=duration,
            elevation_gain=round(elevation_gain, 2),
            elevation_loss=round(elevation_loss, 2),
            start_time=start_time,
            end_time=end_time,
            has_area_info=has_area_info,
            has_road_info=has_road_info,
            created_by=user.id,
            updated_by=user.id,
            is_valid=True,
        )
        db.add(track_obj)
        await db.flush()  # 获取 track_id

        # 批量创建轨迹点
        insert_values = []
        for point_data in points_data:
            insert_values.append({
                "track_id": track_obj.id,
                "point_index": point_data["point_index"],
                "time": point_data["time"],
                "latitude_wgs84": point_data["latitude_wgs84"],
                "longitude_wgs84": point_data["longitude_wgs84"],
                "latitude_gcj02": point_data["latitude_gcj02"],
                "longitude_gcj02": point_data["longitude_gcj02"],
                "latitude_bd09": point_data["latitude_bd09"],
                "longitude_bd09": point_data["longitude_bd09"],
                "elevation": point_data["elevation"],
                "speed": point_data["speed"],
                "bearing": point_data["bearing"],
                "province": point_data.get("province"),
                "city": point_data.get("city"),
                "district": point_data.get("district"),
                "province_en": point_data.get("province_en"),
                "city_en": point_data.get("city_en"),
                "district_en": point_data.get("district_en"),
                "road_name": point_data.get("road_name"),
                "road_number": point_data.get("road_number"),
                "road_name_en": point_data.get("road_name_en"),
                "created_by": user.id,
                "updated_by": user.id,
                "is_valid": True,
            })

        # 分批插入，每批 500 条
        batch_size = 500
        for i in range(0, len(insert_values), batch_size):
            batch = insert_values[i:i + batch_size]
            try:
                await db.execute(
                    TrackPoint.__table__.insert().values(batch)
                )
            except Exception as e:
                logger.error(f"Error inserting batch {i//batch_size}: {e}")
                raise

        await db.commit()
        await db.refresh(track_obj)

        return track_obj

    def _parse_float(self, value: Optional[str]) -> Optional[float]:
        """安全地解析浮点数"""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    async def create_from_kml(
        self,
        db: AsyncSession,
        user: User,
        filename: str,
        kml_content: str,
        name: str,
        description: Optional[str] = None,
        original_crs: CoordinateType = 'wgs84',
        convert_to: Optional[CoordinateType] = None,
        fill_geocoding: bool = False,
    ) -> Track:
        """
        从两步路 KML 内容创建轨迹

        支持的 KML 格式（两步路使用的 Google gx:Track 扩展）：
        - gx:coord: 经度 纬度 海拔
        - when: ISO 8601 格式时间

        Args:
            db: 数据库会话
            user: 用户对象
            filename: 原始文件名
            kml_content: KML 文件内容
            name: 轨迹名称
            description: 轨迹描述
            original_crs: 原始坐标系
            convert_to: 转换到目标坐标系
            fill_geocoding: 是否填充行政区划和道路信息

        Returns:
            创建的轨迹对象
        """
        from lxml import etree

        # 解析 KML
        try:
            root = etree.fromstring(kml_content.encode('utf-8'))
        except etree.ParseError as e:
            raise ValueError(f"无法解析 KML 文件: {e}")

        # 定义命名空间
        ns = {
            'kml': 'http://www.opengis.net/kml/2.2',
            'gx': 'http://www.google.com/kml/ext/2.2'
        }

        # 查找 gx:Track 元素（两步路"轨迹"格式，带时间）
        tracks = root.xpath('.//gx:Track', namespaces=ns)
        if tracks:
            # 使用第一个轨迹
            track_element = tracks[0]

            # 提取坐标和时间
            coord_elements = track_element.xpath('gx:coord', namespaces=ns)
            when_elements = track_element.xpath('kml:when', namespaces=ns)

            if not coord_elements:
                raise ValueError("KML 文件中没有坐标数据")

            # 配对坐标和时间
            # 两者按索引顺序配对，取较小值防止索引越界
            count = min(len(coord_elements), len(when_elements))
        else:
            # 检查是否是两步路"路径"格式（LineString，无时间）
            linestrings = root.xpath('.//kml:LineString', namespaces=ns)
            if linestrings:
                raise ValueError(
                    "检测到这是两步路的“路径”格式（无时间信息）。"
                    "请在两步路应用中导出时选择“KML 格式（轨迹）”，而不是“KML 格式（路径）”。"
                    "轨迹格式包含完整的时间信息，可用于记录运动轨迹。"
                )

            # 检查是否有标准 KML LineString（非两步路）
            linestrings = root.xpath('.//LineString')
            if linestrings:
                raise ValueError(
                    "KML 文件使用 LineString 格式，该格式不包含时间信息。"
                    "请使用包含时间数据的轨迹文件（如 GPX 或两步路“KML 格式（轨迹）”格式）。"
                )

            raise ValueError("KML 文件中没有找到轨迹数据。支持格式：两步路“KML 格式（轨迹）”(KML 中需包含 gx:Track)、GPX")

        # 计算统计信息
        total_distance = 0
        elevation_gain = 0
        elevation_loss = 0
        prev_elevation = None
        start_time = None
        end_time = None

        points_data = []
        prev_point_data = None

        for idx in range(count):
            # 解析坐标 (格式: longitude latitude elevation)
            coord_text = coord_elements[idx].text.strip()
            if not coord_text:
                continue

            coord_parts = coord_text.split()
            if len(coord_parts) < 2:
                logger.warning(f"跳过无效坐标的行 {idx}: {coord_text}")
                continue

            try:
                lon = float(coord_parts[0])
                lat = float(coord_parts[1])
                elevation = float(coord_parts[2]) if len(coord_parts) > 2 else None
            except ValueError as e:
                logger.warning(f"跳过无效坐标的行 {idx}: {e}")
                continue

            # 解析时间
            point_time = None
            if idx < len(when_elements):
                time_str = when_elements[idx].text.strip()
                if time_str:
                    try:
                        # ISO 8601 格式，如 2026-01-09T06:53:49Z
                        if time_str.endswith('Z'):
                            time_str = time_str[:-1] + '+00:00'
                        point_time = datetime.fromisoformat(time_str)
                        # 移除时区信息（与数据库保持一致）
                        point_time = point_time.replace(tzinfo=None)

                        if start_time is None:
                            start_time = point_time
                        end_time = point_time
                    except ValueError as e:
                        logger.warning(f"无法解析时间 '{time_str}': {e}")

            # 坐标转换
            all_coords = convert_point_to_all(lon, lat, original_crs)

            # 海拔变化
            if elevation is not None and prev_elevation is not None:
                diff = elevation - prev_elevation
                if diff > 0:
                    elevation_gain += diff
                else:
                    elevation_loss += abs(diff)
            prev_elevation = elevation

            # 计算距离和速度
            distance_from_prev = 0
            speed = None
            if idx > 0 and prev_point_data:
                # 使用 WGS84 坐标计算 3D 距离
                from math import sqrt, radians, sin, cos, atan2

                lat1 = radians(prev_point_data['latitude_wgs84'])
                lon1 = radians(prev_point_data['longitude_wgs84'])
                lat2 = radians(all_coords['wgs84'][1])
                lon2 = radians(all_coords['wgs84'][0])

                # Haversine 公式计算水平距离
                dlat = lat2 - lat1
                dlon = lon2 - lon1
                a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
                c = 2 * atan2(sqrt(a), sqrt(1 - a))
                horizontal_distance = 6371000 * c  # 地球半径 6371km

                # 计算垂直距离
                elev1 = prev_point_data.get('elevation') or 0
                elev2 = elevation or 0
                vertical_distance = elev2 - elev1

                # 3D 距离
                distance_from_prev = sqrt(horizontal_distance ** 2 + vertical_distance ** 2)
                total_distance += distance_from_prev

                # 计算速度
                speed = self._calculate_speed(point_time, prev_point_data['time'], distance_from_prev)

                # 计算方位角
                bearing = self._calculate_bearing(
                    prev_point_data['latitude_wgs84'],
                    prev_point_data['longitude_wgs84'],
                    all_coords['wgs84'][1],
                    all_coords['wgs84'][0]
                )

            point_data = {
                'point_index': idx,
                'time': point_time,
                'latitude_wgs84': all_coords['wgs84'][1],
                'longitude_wgs84': all_coords['wgs84'][0],
                'latitude_gcj02': all_coords['gcj02'][1],
                'longitude_gcj02': all_coords['gcj02'][0],
                'latitude_bd09': all_coords['bd09'][1],
                'longitude_bd09': all_coords['bd09'][0],
                'elevation': elevation,
                'speed': speed,
                'bearing': bearing if idx > 0 else None,  # 第一个点没有方位角
            }
            points_data.append(point_data)
            prev_point_data = point_data

        if not points_data:
            raise ValueError("KML 文件中没有有效的轨迹点")

        # 计算时长（秒）
        duration = 0
        if start_time and end_time:
            duration = int((end_time - start_time).total_seconds())

        # 创建轨迹记录，带审计字段
        track_obj = Track(
            user_id=user.id,
            name=name,
            description=description,
            original_filename=filename,
            original_crs=original_crs,
            distance=round(total_distance, 2),
            duration=duration,
            elevation_gain=round(elevation_gain, 2),
            elevation_loss=round(elevation_loss, 2),
            start_time=start_time,
            end_time=end_time,
            has_area_info=False,
            has_road_info=False,
            created_by=user.id,
            updated_by=user.id,
            is_valid=True,
        )
        db.add(track_obj)
        await db.flush()  # 获取 track_id

        # 批量创建轨迹点
        insert_values = []
        for point_data in points_data:
            insert_values.append({
                "track_id": track_obj.id,
                "point_index": point_data["point_index"],
                "time": point_data["time"],
                "latitude_wgs84": point_data["latitude_wgs84"],
                "longitude_wgs84": point_data["longitude_wgs84"],
                "latitude_gcj02": point_data["latitude_gcj02"],
                "longitude_gcj02": point_data["longitude_gcj02"],
                "latitude_bd09": point_data["latitude_bd09"],
                "longitude_bd09": point_data["longitude_bd09"],
                "elevation": point_data["elevation"],
                "speed": point_data["speed"],
                "bearing": point_data["bearing"],
                "province": None,
                "city": None,
                "district": None,
                "province_en": None,
                "city_en": None,
                "district_en": None,
                "road_name": None,
                "road_number": None,
                "road_name_en": None,
                "created_by": user.id,
                "updated_by": user.id,
                "is_valid": True,
            })

        # 分批插入，每批 500 条
        batch_size = 500
        for i in range(0, len(insert_values), batch_size):
            batch = insert_values[i:i + batch_size]
            try:
                await db.execute(
                    TrackPoint.__table__.insert().values(batch)
                )
            except Exception as e:
                logger.error(f"Error inserting batch {i//batch_size}: {e}")
                raise

        await db.commit()
        await db.refresh(track_obj)

        # 异步填充地理信息（如果启用）
        if fill_geocoding:
            task = asyncio.create_task(
                self.fill_geocoding_info(db, track_obj.id, user.id)
            )
            task.add_done_callback(self._background_tasks.discard)
            self._background_tasks.add(task)

        return track_obj

    async def create_from_xlsx(
        self,
        db: AsyncSession,
        user: User,
        filename: str,
        xlsx_content: bytes,
        name: str,
        description: Optional[str] = None,
    ) -> Track:
        """
        从本项目导出的 XLSX 文件创建轨迹

        支持的字段：
        - index, time_date, time_time, time_microsecond, elapsed_time
        - longitude_wgs84, latitude_wgs84
        - longitude_gcj02, latitude_gcj02
        - longitude_bd09, latitude_bd09
        - elevation, distance, course, speed
        - province, city, area, province_en, city_en, area_en
        - road_num, road_name, road_name_en, memo

        Args:
            db: 数据库会话
            user: 用户对象
            filename: 原始文件名
            xlsx_content: XLSX 文件二进制内容
            name: 轨迹名称
            description: 轨迹描述

        Returns:
            创建的轨迹对象
        """
        from openpyxl import load_workbook
        from io import BytesIO

        # 读取 XLSX 文件
        wb = load_workbook(BytesIO(xlsx_content), read_only=True)
        ws = wb.active

        if ws is None:
            raise ValueError("XLSX 文件格式错误：没有活动工作表")

        # 获取表头
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        if not headers or headers[0] is None:
            raise ValueError("XLSX 文件格式错误：缺少表头")

        # 检查是否是本项目导出格式
        if 'longitude_wgs84' not in headers or 'latitude_wgs84' not in headers:
            raise ValueError("XLSX 文件格式不支持：缺少必需的坐标字段 (longitude_wgs84, latitude_wgs84)")

        # 构建字典列表（类似 CSV DictReader 的格式）
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue

            row_dict = {}
            for col_idx, header in enumerate(headers):
                if header and col_idx < len(row):
                    row_dict[header] = str(row[col_idx]) if row[col_idx] is not None else ""
            rows.append(row_dict)

        if not rows:
            raise ValueError("XLSX 文件中没有数据")

        # 使用与 CSV 相同的解析逻辑
        return await self._create_from_csv_project_format(
            db, user, filename, rows, name, description
        )

    async def change_original_crs(
        self,
        db: AsyncSession,
        track_id: int,
        user_id: int,
        new_original_crs: CoordinateType,
    ) -> Track:
        """
        更改轨迹的原始坐标系，并重新计算所有坐标系

        场景：用户上传时选错了坐标系，现在要更正。
        例如：上传时选了 WGS84，但实际数据是 GCJ02。

        处理逻辑：
        - 从旧的 original_crs 对应字段读取坐标（这些数据实际是新坐标系的坐标）
        - 使用新坐标系重新计算所有坐标系的坐标

        Args:
            db: 数据库会话
            track_id: 轨迹 ID
            user_id: 用户 ID
            new_original_crs: 新的原始坐标系（实际数据的坐标系）

        Returns:
            更新后的轨迹对象
        """
        # 获取轨迹
        track = await self.get_by_id(db, track_id, user_id)
        if not track:
            raise ValueError("轨迹不存在")

        old_original_crs = track.original_crs

        # 如果坐标系相同，直接返回
        if old_original_crs == new_original_crs:
            return track

        # 获取所有轨迹点
        result = await db.execute(
            select(TrackPoint)
            .where(and_(TrackPoint.track_id == track_id, TrackPoint.is_valid == True))
            .order_by(TrackPoint.point_index)
        )
        points = list(result.scalars().all())

        if not points:
            raise ValueError("轨迹没有数据点")

        # 重新计算所有坐标系
        # 从旧的 original_crs 对应字段读取坐标，这些数据实际是新坐标系的坐标
        for point in points:
            # 从旧坐标系字段读取原始坐标（这些数据实际是新坐标系的坐标）
            if old_original_crs == 'wgs84':
                actual_lon = point.longitude_wgs84
                actual_lat = point.latitude_wgs84
            elif old_original_crs == 'gcj02':
                actual_lon = point.longitude_gcj02
                actual_lat = point.latitude_gcj02
            elif old_original_crs == 'bd09':
                actual_lon = point.longitude_bd09
                actual_lat = point.latitude_bd09
            else:
                raise ValueError(f"不支持的坐标系: {old_original_crs}")

            if actual_lon is None or actual_lat is None:
                logger.warning(f"点 {point.point_index} 缺少坐标数据，跳过")
                continue

            # 使用新坐标系重新计算所有坐标系
            all_coords = convert_point_to_all(actual_lon, actual_lat, new_original_crs)

            # 更新坐标
            point.longitude_wgs84 = all_coords['wgs84'][0]
            point.latitude_wgs84 = all_coords['wgs84'][1]
            point.longitude_gcj02 = all_coords['gcj02'][0]
            point.latitude_gcj02 = all_coords['gcj02'][1]
            point.longitude_bd09 = all_coords['bd09'][0]
            point.latitude_bd09 = all_coords['bd09'][1]

            point.updated_by = user_id

        # 更新轨迹的原始坐标系
        track.original_crs = new_original_crs
        track.updated_by = user_id

        await db.commit()
        await db.refresh(track)

        logger.info(f"Track {track_id}: changed original_crs to {new_original_crs}, updated {len(points)} points")

        return track


track_service = TrackService()
